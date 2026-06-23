"""Import staff, locations, and schedule data from the existing Excel file."""

import re
from openpyxl import load_workbook
from . import database as db
from .config import FIXED_ASSIGNMENTS, FIXED_ROLES, LOCATION_CODE_MAP


# Map Excel shift strings to canonical labels
SHIFT_NORMALIZE = {
    "1-10": "1-10",
    "2-11": "2-11",
    "3-12": "3-12",
    "3.12": "3-12",
    "10-7": "10-7",
    "11-30/8.30": "11:30-8:30",
    "11-30/8:30": "11:30-8:30",
    "12-9": "12-9",
}


def parse_cell(value: str):
    """Parse a cell value like '10-7 MGM' into (shift_label, location_code, status)."""
    if not value or not value.strip():
        return None, None, "off"

    value = value.strip()
    upper = value.upper()

    if upper == "OFF":
        return None, None, "off"
    elif upper == "LEAVE":
        return None, None, "leave"
    elif upper in ("SALALAH", "NIZWA", "CLINIC", "OFFICE"):
        # Fixed role/location — no specific shift
        code_map = {"SALALAH": "SALALAH", "NIZWA": "NIZWA", "CLINIC": "Clinic", "OFFICE": "Office"}
        return None, code_map[upper], "assigned"

    # Try to parse "shift location" format like "10-7 MGM" or "1-10 AV"
    parts = value.split()
    if len(parts) >= 2:
        shift_raw = parts[0]
        loc_raw = parts[1].upper()
        shift_label = SHIFT_NORMALIZE.get(shift_raw, shift_raw)
        loc_code = loc_raw
        # Map common abbreviations
        if loc_code == "AV":
            loc_code = "AV"
        return shift_label, loc_code, "assigned"

    # Single value that might be a location
    if upper in LOCATION_CODE_MAP or upper in [v.upper() for v in LOCATION_CODE_MAP.values()]:
        return None, value, "assigned"

    return None, None, "off"


def import_excel(filepath: str):
    """Import data from the Excel schedule file."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    # Row 2 has staff names (columns B-Y = indices 2-25)
    staff_names = []
    for col in range(2, ws.max_column + 1):
        name = ws.cell(row=2, column=col).value
        if name and str(name).strip():
            staff_names.append((col, str(name).strip()))

    # Add all staff to database
    for _, name in staff_names:
        default_loc = FIXED_ASSIGNMENTS.get(name)
        default_role = FIXED_ROLES.get(name)
        try:
            db.add_staff(name, default_location=default_loc, default_role=default_role)
        except Exception:
            pass  # Already exists

    # Schedule dates are in rows 3-9, column A
    # Row 3: Sat 28/3, Row 4: Sun 29/3, etc.
    day_rows = list(range(3, 10))  # rows 3-9 = 7 days

    # Create a schedule for this week
    schedule_id = db.create_schedule("2026-03-28", "2026-04-03", "Imported from Excel")

    all_staff = {s.name: s for s in db.get_all_staff(active_only=False)}
    all_locations = {l.short_code: l for l in db.get_all_locations(active_only=False)}
    all_shifts = {s.label: s for s in db.get_all_shifts(active_only=False)}

    # Date mapping for each row
    dates = ["2026-03-28", "2026-03-29", "2026-03-30", "2026-03-31",
             "2026-04-01", "2026-04-02", "2026-04-03"]

    for day_idx, row_num in enumerate(day_rows):
        for col, name in staff_names:
            cell_value = ws.cell(row=row_num, column=col).value
            cell_str = str(cell_value).strip() if cell_value else ""

            shift_label, loc_code, status = parse_cell(cell_str)

            staff = all_staff.get(name)
            if not staff:
                continue

            location_id = None
            shift_id = None

            if loc_code and loc_code in all_locations:
                location_id = all_locations[loc_code].id
            elif loc_code:
                # Try case-insensitive match
                for code, loc in all_locations.items():
                    if code.upper() == loc_code.upper() or loc.name.upper() == loc_code.upper():
                        location_id = loc.id
                        break

            if shift_label and shift_label in all_shifts:
                shift_id = all_shifts[shift_label].id
            elif shift_label:
                # Add new shift if not found
                new_id = db.add_shift(shift_label, "", "")
                if new_id:
                    shift_id = new_id
                    all_shifts[shift_label] = db.get_shift_by_label(shift_label)

            db.add_assignment(
                schedule_id=schedule_id,
                staff_id=staff.id,
                day_of_week=day_idx,
                day_date=dates[day_idx],
                status=status,
                location_id=location_id,
                shift_id=shift_id
            )

    wb.close()
    return schedule_id

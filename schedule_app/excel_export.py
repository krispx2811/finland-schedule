"""Excel export using openpyxl."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from . import database as db
from .config import DAYS_SHORT

STATUS_FILLS = {
    "off": PatternFill(start_color="D5D8DC", fill_type="solid"),
    "leave": PatternFill(start_color="F9E79F", fill_type="solid"),
}

LOCATION_FILLS = {
    "MOO": PatternFill(start_color="991B1B", fill_type="solid"),       # Dark Red
    "MGM": PatternFill(start_color="FFF200", fill_type="solid"),       # Neon Yellow
    "SCC": PatternFill(start_color="831843", fill_type="solid"),       # Dark Dark Pink
    "AV": PatternFill(start_color="FED7AA", fill_type="solid"),        # Orange-Brown
    "CCC": PatternFill(start_color="D1FAE5", fill_type="solid"),       # Green
    "QCC": PatternFill(start_color="DBEAFE", fill_type="solid"),       # Blue
    "SALALAH": PatternFill(start_color="C8940A", fill_type="solid"),   # Gold
    "NIZWA": PatternFill(start_color="E9D5FF", fill_type="solid"),     # Purple
    "Clinic": PatternFill(start_color="DBEAFE", fill_type="solid"),    # Blue
    "Office": PatternFill(start_color="D1FAE5", fill_type="solid"),    # Green
}


def export_schedule_excel(schedule_id, filepath):
    schedule = None
    for s in db.get_all_schedules():
        if s.id == schedule_id:
            schedule = s
            break
    if not schedule:
        return

    assignments = db.get_schedule_assignments(schedule_id)
    staff_data = {}
    for a in assignments:
        if a.staff_name not in staff_data:
            staff_data[a.staff_name] = [""] * 7
        staff_data[a.staff_name][a.day_of_week] = a

    wb = Workbook()
    ws = wb.active
    ws.title = f"Week {schedule.week_start}"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A1A2E", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Title row
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Finland Optical Center — {schedule.week_start} to {schedule.week_end}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header row
    headers = ["Staff"] + DAYS_SHORT
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for row_idx, name in enumerate(sorted(staff_data.keys()), 4):
        ws.cell(row=row_idx, column=1, value=name).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).border = thin_border

        for col_idx in range(7):
            a = staff_data[name][col_idx]
            cell = ws.cell(row=row_idx, column=col_idx + 2)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

            if isinstance(a, str):
                cell.value = a
                continue

            text = a.display_text()
            cell.value = text

            upper = text.upper()
            if upper == "OFF":
                cell.fill = STATUS_FILLS["off"]
            elif upper == "LEAVE":
                cell.fill = STATUS_FILLS["leave"]
            else:
                dark_bg_codes = {"MOO", "SCC", "SALALAH"}
                for code, fill in LOCATION_FILLS.items():
                    if code.upper() in upper:
                        cell.fill = fill
                        if code.upper() in dark_bg_codes:
                            cell.font = Font(color="FFFFFF", size=11)
                        break

    # Column widths
    ws.column_dimensions['A'].width = 18
    for col in 'BCDEFGH':
        ws.column_dimensions[col].width = 16

    # Freeze panes
    ws.freeze_panes = 'B4'

    wb.save(filepath)
